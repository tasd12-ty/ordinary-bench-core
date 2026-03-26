#!/usr/bin/env python3
"""
Export evaluation results to Excel and run batch belief reconstruction with SVG.

Usage:
    python analysis/export_and_reconstruct.py                         # both
    python analysis/export_and_reconstruct.py --excel-only            # Excel only
    python analysis/export_and_reconstruct.py --recon-only            # recon + SVG only
    python analysis/export_and_reconstruct.py --models claude-sonnet-4-6 --max-scenes 3
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np

# Ensure imports work from VLM-test/
sys.path.insert(0, str(Path(__file__).parent.parent))

from analysis.aggregate import load_scene_results
from analysis.reconstruct_scenes import reconstruct_single_scene
from analysis.visualize_svg import (
    load_object_info,
    render_scene_comparison_svg,
    ObjectInfo,
)
from reconstruct import load_questions_auto, load_scene_gt_positions


# ── Paths ──

BASE_DIR = Path(__file__).parent.parent  # VLM-test/
OUTPUT_DIR = BASE_DIR / "output"
RESULTS_DIR = OUTPUT_DIR / "results"
QUESTIONS_DIR = OUTPUT_DIR / "questions"
SCENES_DIR = BASE_DIR.parent / "data-gen" / "output" / "scenes"
ANALYSIS_DIR = OUTPUT_DIR / "analysis"
RECON_DIR = ANALYSIS_DIR / "belief_recon"


def detect_view_type(model_dir_name: str) -> str:
    return "multi" if "multi_view" in model_dir_name else "single"


def discover_models(filter_models: Optional[List[str]] = None) -> List[Tuple[str, Path]]:
    """Return sorted list of (model_dir_name, path) for models with scene results."""
    models = []
    for d in sorted(RESULTS_DIR.iterdir()):
        if not d.is_dir():
            continue
        if not (d / "scenes").is_dir():
            continue
        if filter_models and d.name not in filter_models:
            continue
        models.append((d.name, d))
    return models


# ── Part 1: Excel Export ──

def export_excel(models: List[Tuple[str, Path]], output_path: Path):
    """Create Excel workbook with Overall, By Split, and Per Scene sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers

    wb = Workbook()

    # ── Sheet: Overall ──
    ws_overall = wb.active
    ws_overall.title = "Overall"

    overall_headers = [
        "Model", "View", "N_Scenes",
        "QRR_Acc", "QRR_Disjoint_Acc", "QRR_SharedAnchor_Acc",
        "TRR_Hour_Acc", "TRR_Quadrant_Acc", "TRR_Adjacent_Acc",
        "FDR_Exact_Acc", "FDR_Kendall_Mean", "FDR_Pairwise_Mean", "FDR_Top1_Mean",
        "Missing_Count",
    ]
    ws_overall.append(overall_headers)

    # Style header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for cell in ws_overall[1]:
        cell.font = header_font
        cell.fill = header_fill

    for model_name, model_path in models:
        summary_path = model_path / "summary.json"
        if not summary_path.exists():
            continue
        with open(summary_path) as f:
            summary = json.load(f)
        o = summary.get("overall", {})
        view = detect_view_type(model_name)
        ws_overall.append([
            model_name, view, summary.get("n_scenes", 0),
            o.get("qrr_accuracy"),
            o.get("qrr_disjoint_accuracy"),
            o.get("qrr_shared_anchor_accuracy"),
            o.get("trr_hour_accuracy"),
            o.get("trr_quadrant_accuracy"),
            o.get("trr_adjacent_accuracy"),
            o.get("fdr_exact_accuracy"),
            o.get("fdr_kendall_mean"),
            o.get("fdr_pairwise_mean"),
            o.get("fdr_top1_mean"),
            o.get("missing", 0),
        ])

    # Format percentage columns
    for row in ws_overall.iter_rows(min_row=2, min_col=4, max_col=13):
        for cell in row:
            if cell.value is not None:
                cell.number_format = '0.00%'

    # Auto-width
    for col in ws_overall.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_overall.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

    # ── Sheet: By Split ──
    ws_split = wb.create_sheet("By Split")
    split_headers = [
        "Model", "View", "Split", "N_Scenes",
        "QRR_Acc", "TRR_Hour_Acc", "TRR_Quadrant_Acc", "TRR_Adjacent_Acc",
        "FDR_Exact_Acc", "FDR_Kendall_Mean", "FDR_Pairwise_Mean", "FDR_Top1_Mean",
        "Missing_Count",
    ]
    ws_split.append(split_headers)
    for cell in ws_split[1]:
        cell.font = header_font
        cell.fill = header_fill

    for model_name, model_path in models:
        summary_path = model_path / "summary.json"
        if not summary_path.exists():
            continue
        with open(summary_path) as f:
            summary = json.load(f)
        view = detect_view_type(model_name)
        by_split = summary.get("by_split", {})
        for split_name in sorted(by_split.keys()):
            s = by_split[split_name]
            # Count scenes in this split
            n_scenes_in_split = 0
            scenes_dir = model_path / "scenes"
            if scenes_dir.exists():
                n_scenes_in_split = sum(
                    1 for f in scenes_dir.glob(f"{split_name}_*.json")
                )
            ws_split.append([
                model_name, view, split_name, n_scenes_in_split,
                s.get("qrr_accuracy"),
                s.get("trr_hour_accuracy"),
                s.get("trr_quadrant_accuracy"),
                s.get("trr_adjacent_accuracy"),
                s.get("fdr_exact_accuracy"),
                s.get("fdr_kendall_mean"),
                s.get("fdr_pairwise_mean"),
                s.get("fdr_top1_mean"),
                s.get("missing", 0),
            ])

    for row in ws_split.iter_rows(min_row=2, min_col=5, max_col=12):
        for cell in row:
            if cell.value is not None:
                cell.number_format = '0.00%'

    for col in ws_split.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_split.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

    # ── Sheet: Per Scene ──
    ws_scene = wb.create_sheet("Per Scene")
    scene_headers = [
        "Model", "View", "Scene_ID", "Split", "N_Objects",
        "QRR_Correct", "QRR_Total", "QRR_Acc",
        "TRR_Hour_Correct", "TRR_Quadrant_Correct", "TRR_Adjacent_Correct", "TRR_Total",
        "TRR_Hour_Acc", "TRR_Quadrant_Acc", "TRR_Adjacent_Acc",
        "FDR_Exact_Correct", "FDR_Total", "FDR_Exact_Acc",
        "FDR_Kendall_Mean", "FDR_Pairwise_Mean", "FDR_Top1_Mean",
        "Missing",
    ]
    ws_scene.append(scene_headers)
    for cell in ws_scene[1]:
        cell.font = header_font
        cell.fill = header_fill

    for model_name, model_path in models:
        view = detect_view_type(model_name)
        scene_results = load_scene_results(str(model_path))
        for sr in scene_results:
            scene_id = sr["scene_id"]
            split = scene_id.rsplit("_", 1)[0]
            n_obj = sr.get("n_objects", int(scene_id[1:3]))
            sc = sr["scores"]

            def _safe_div(a, b):
                return a / b if b > 0 else None

            ws_scene.append([
                model_name, view, scene_id, split, n_obj,
                sc.get("qrr_correct", 0), sc.get("qrr_total", 0),
                _safe_div(sc.get("qrr_correct", 0), sc.get("qrr_total", 0)),
                sc.get("trr_hour_correct", 0),
                sc.get("trr_quadrant_correct", 0),
                sc.get("trr_adjacent_correct", 0),
                sc.get("trr_total", 0),
                _safe_div(sc.get("trr_hour_correct", 0), sc.get("trr_total", 0)),
                _safe_div(sc.get("trr_quadrant_correct", 0), sc.get("trr_total", 0)),
                _safe_div(sc.get("trr_adjacent_correct", 0), sc.get("trr_total", 0)),
                sc.get("fdr_exact_correct", 0), sc.get("fdr_total", 0),
                _safe_div(sc.get("fdr_exact_correct", 0), sc.get("fdr_total", 0)),
                sc.get("fdr_kendall_mean"),
                sc.get("fdr_pairwise_mean"),
                sc.get("fdr_top1_mean"),
                sc.get("missing", 0),
            ])

    # Format percentage columns (8, 13-15, 18)
    pct_cols = [8, 13, 14, 15, 18]
    for row in ws_scene.iter_rows(min_row=2):
        for idx in pct_cols:
            cell = row[idx - 1]  # 0-indexed
            if cell.value is not None:
                cell.number_format = '0.00%'

    for col in ws_scene.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_scene.column_dimensions[col[0].column_letter].width = min(max_len + 3, 22)

    # Freeze panes
    ws_overall.freeze_panes = "A2"
    ws_split.freeze_panes = "A2"
    ws_scene.freeze_panes = "A2"

    os.makedirs(output_path.parent, exist_ok=True)
    wb.save(str(output_path))
    print(f"Excel saved: {output_path}")
    return wb


# ── Part 2: Belief Reconstruction + SVG ──

def load_scene_gt(scene_path: str) -> Optional[Dict[str, np.ndarray]]:
    loaded = load_scene_gt_positions(scene_path)
    if not loaded:
        return None
    return {oid: np.array(pos, dtype=np.float64) for oid, pos in loaded.items()}


def run_belief_reconstruction(
    models: List[Tuple[str, Path]],
    max_scenes: Optional[int] = None,
    n_restarts: int = 10,
) -> List[dict]:
    """Run belief reconstruction for all models and generate SVGs.

    Returns list of per-scene recon result dicts (for Excel sheet).
    """
    all_recon_rows = []

    for model_name, model_path in models:
        print(f"\n{'='*60}")
        print(f"Reconstructing: {model_name}")
        print(f"{'='*60}")

        scene_results = load_scene_results(str(model_path))
        if not scene_results:
            print(f"  No scene results found, skipping")
            continue

        if max_scenes:
            scene_results = scene_results[:max_scenes]

        model_recon_dir = RECON_DIR / model_name
        model_recon_dir.mkdir(parents=True, exist_ok=True)

        for i, sr in enumerate(scene_results):
            scene_id = sr["scene_id"]

            # Load questions
            questions, question_meta = load_questions_auto(str(QUESTIONS_DIR), scene_id)
            if not questions:
                print(f"  [{i+1}/{len(scene_results)}] {scene_id}: no questions, skip")
                continue

            # Load GT
            scene_path = str(SCENES_DIR / f"{scene_id}.json")
            gt_positions = load_scene_gt(scene_path) if os.path.exists(scene_path) else None

            # Reconstruct (belief mode: use_correct_only=False)
            try:
                recon_output = reconstruct_single_scene(
                    scene_result=sr,
                    questions=questions,
                    gt_positions=gt_positions,
                    use_correct_only=False,
                    n_restarts=n_restarts,
                    question_metadata=question_meta,
                )

                status = recon_output["status"]
                m = recon_output["metrics"]
                nrms_val = m.get("nrms")
                nrms_str = f"{nrms_val:.4f}" if isinstance(nrms_val, (int, float)) else "N/A"
                print(
                    f"  [{i+1}/{len(scene_results)}] {scene_id}: "
                    f"status={status} csr_qrr={m['csr_qrr']:.3f} "
                    f"csr_trr={m['csr_trr']:.3f} nrms={nrms_str}"
                )

                # Save recon JSON
                recon_json_path = model_recon_dir / f"{scene_id}.json"
                with open(recon_json_path, "w") as f:
                    json.dump(recon_output, f, indent=2, default=str)

                # Generate SVG if we have positions and GT
                if recon_output.get("positions") and gt_positions and os.path.exists(scene_path):
                    try:
                        recon_positions = {
                            oid: np.array(coords, dtype=float)
                            for oid, coords in recon_output["positions"].items()
                        }
                        gt_2d = {oid: pos[:2] for oid, pos in gt_positions.items()}
                        obj_info = load_object_info(scene_path)

                        svg_str = render_scene_comparison_svg(
                            gt_positions=gt_2d,
                            recon_positions=recon_positions,
                            object_info=obj_info,
                            scene_id=f"{scene_id} | {model_name}",
                            metrics=recon_output["metrics"],
                            panel_size=320,
                        )

                        svg_path = model_recon_dir / f"{scene_id}.svg"
                        with open(svg_path, "w") as f:
                            f.write(svg_str)
                    except Exception as svg_err:
                        print(f"    SVG generation failed: {svg_err}")

                # Collect for Excel
                all_recon_rows.append({
                    "model": model_name,
                    "scene_id": scene_id,
                    "n_objects": sr.get("n_objects", 0),
                    "status": status,
                    "feasible": recon_output.get("feasible", False),
                    "csr_qrr": m.get("csr_qrr"),
                    "csr_trr": m.get("csr_trr"),
                    "nrms": m.get("nrms"),
                    "kendall_tau": m.get("kendall_tau"),
                    "K_geom": m.get("K_geom"),
                    "best_loss": m.get("best_loss"),
                    "n_solutions": m.get("n_solutions"),
                    "reflected": m.get("reflected", False),
                })

            except Exception as e:
                print(f"  [{i+1}/{len(scene_results)}] {scene_id}: ERROR {e}")
                all_recon_rows.append({
                    "model": model_name,
                    "scene_id": scene_id,
                    "n_objects": sr.get("n_objects", 0),
                    "status": "error",
                    "feasible": False,
                    "error": str(e),
                })

    return all_recon_rows


def add_recon_sheet(wb, recon_rows: List[dict]):
    """Add Reconstruction sheet to existing workbook."""
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Reconstruction")
    headers = [
        "Model", "Scene_ID", "N_Objects", "Status", "Feasible",
        "CSR_QRR", "CSR_TRR", "NRMS", "Kendall_Tau", "K_geom",
        "Best_Loss", "N_Solutions", "Reflected",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in recon_rows:
        ws.append([
            r.get("model"),
            r.get("scene_id"),
            r.get("n_objects"),
            r.get("status"),
            r.get("feasible"),
            r.get("csr_qrr"),
            r.get("csr_trr"),
            r.get("nrms"),
            r.get("kendall_tau"),
            r.get("K_geom"),
            r.get("best_loss"),
            r.get("n_solutions"),
            r.get("reflected", False),
        ])

    # Format
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=9):
        for cell in row:
            if cell.value is not None:
                cell.number_format = '0.0000'

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 22)

    ws.freeze_panes = "A2"


# ── Main ──

def main():
    parser = argparse.ArgumentParser(description="Export results to Excel and run belief reconstruction")
    parser.add_argument("--excel-only", action="store_true", help="Only export Excel")
    parser.add_argument("--recon-only", action="store_true", help="Only run reconstruction + SVG")
    parser.add_argument("--models", type=str, default=None,
                        help="Comma-separated model dir names to process")
    parser.add_argument("--max-scenes", type=int, default=None,
                        help="Max scenes per model for reconstruction")
    parser.add_argument("--restarts", type=int, default=10,
                        help="Number of optimizer restarts")
    parser.add_argument("--output", type=str, default=None,
                        help="Excel output path (default: output/analysis/results_summary.xlsx)")
    args = parser.parse_args()

    filter_models = args.models.split(",") if args.models else None
    models = discover_models(filter_models)
    print(f"Found {len(models)} models: {[m[0] for m in models]}")

    excel_path = Path(args.output) if args.output else ANALYSIS_DIR / "results_summary.xlsx"

    if not args.recon_only:
        # Part 1: Excel
        wb = export_excel(models, excel_path)

        if not args.excel_only:
            # Part 2: Reconstruction
            recon_rows = run_belief_reconstruction(
                models, max_scenes=args.max_scenes, n_restarts=args.restarts,
            )
            if recon_rows:
                add_recon_sheet(wb, recon_rows)
                wb.save(str(excel_path))
                print(f"\nUpdated Excel with Reconstruction sheet: {excel_path}")
    else:
        # Recon only — load or create workbook
        from openpyxl import load_workbook, Workbook
        if excel_path.exists():
            wb = load_workbook(str(excel_path))
        else:
            wb = Workbook()

        recon_rows = run_belief_reconstruction(
            models, max_scenes=args.max_scenes, n_restarts=args.restarts,
        )
        if recon_rows:
            # Remove old Reconstruction sheet if exists
            if "Reconstruction" in wb.sheetnames:
                del wb["Reconstruction"]
            add_recon_sheet(wb, recon_rows)
            wb.save(str(excel_path))
            print(f"\nSaved Reconstruction sheet to: {excel_path}")

    print("\nDone.")


if __name__ == "__main__":
    main()
