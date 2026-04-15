#!/usr/bin/env python3
"""
Build pivot-style Excel summaries for quick-sort BT=0 reconstruction outputs.

This script expects:
1. a reconstruction root containing per-run `recon/summary.json`
2. the original quick-sort source results tree containing `summary.json + scenes/`

It writes a workbook with:
- scene_long
- pivot_model_run_n
- pivot_model_run
- pivot_model_n
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Optional


KNOWN_STATUS_COLUMNS = ("single_mode", "multimodal", "infeasible")
MEAN_METRIC_COLUMNS = (
    "csr_qrr",
    "csr_qrr_aligned",
    "kendall_tau",
    "nrms",
)
SCENE_LONG_COLUMNS = (
    "model_family",
    "raw_model",
    "run_label",
    "source_attempt",
    "canonical_run",
    "scene_id",
    "n_objects",
    "status",
    "feasible",
    "is_skipped",
    "skip_reason",
    "source_scene_result",
    "csr_qrr",
    "csr_qrr_aligned",
    "kendall_tau",
    "nrms",
    "best_loss",
    "n_solutions",
)
PIVOT_MODEL_RUN_N_COLUMNS = (
    "model_family",
    "run_label",
    "n_objects",
    "source_scenes",
    "reconstructed",
    "skipped",
    "feasible",
    "single_mode",
    "multimodal",
    "infeasible",
    "other_status",
    "feasible_reconstructed",
    "feasible_total",
    "csr_qrr_mean",
    "csr_qrr_aligned_mean",
    "kendall_tau_mean",
    "nrms_mean",
)
PIVOT_MODEL_RUN_COLUMNS = (
    "model_family",
    "run_label",
    "source_scenes",
    "reconstructed",
    "skipped",
    "feasible",
    "single_mode",
    "multimodal",
    "infeasible",
    "other_status",
    "feasible_reconstructed",
    "feasible_total",
    "csr_qrr_mean",
    "csr_qrr_aligned_mean",
    "kendall_tau_mean",
    "nrms_mean",
    "status_counts",
)
PIVOT_MODEL_N_COLUMNS = (
    "model_family",
    "n_objects",
    "source_scenes",
    "reconstructed",
    "skipped",
    "feasible",
    "single_mode",
    "multimodal",
    "infeasible",
    "other_status",
    "feasible_reconstructed",
    "feasible_total",
    "csr_qrr_mean",
    "csr_qrr_aligned_mean",
    "kendall_tau_mean",
    "nrms_mean",
)


def _load_json(path: Path) -> dict[str, Any]:
    with open(path) as f:
        return json.load(f)


def _discover_run_dirs(results_root: Path) -> list[Path]:
    if (results_root / "summary.json").is_file() and (results_root / "scenes").is_dir():
        return [results_root]
    run_dirs = sorted(
        {
            summary_path.parent
            for summary_path in results_root.rglob("summary.json")
            if (summary_path.parent / "scenes").is_dir()
        }
    )
    if not run_dirs:
        raise FileNotFoundError(
            f"No run directories found under {results_root} (expected summary.json + scenes/)"
        )
    return run_dirs


def _discover_recon_run_dirs(recon_root: Path) -> list[Path]:
    run_dirs = sorted(
        {
            summary_path.parent.parent
            for summary_path in recon_root.rglob("recon/summary.json")
        }
    )
    if not run_dirs:
        raise FileNotFoundError(
            f"No reconstruction run directories found under {recon_root}"
        )
    return run_dirs


def normalize_model_family(raw_model: Optional[str]) -> str:
    if not raw_model:
        return "unknown"
    raw = raw_model.lower()
    if "qwen" not in raw:
        return raw_model
    if "397" in raw:
        return "qwen397"
    if re.search(r"(^|[-_])27($|[^0-9])", raw):
        return "qwen27"
    if re.search(r"(^|[-_])9($|[^0-9])", raw):
        return "qwen9"
    return raw_model


def infer_n_objects(scene_id: str, declared: Optional[int] = None) -> int:
    if declared is not None:
        return int(declared)
    match = re.search(r"__sz(\d+)_", scene_id)
    if match:
        return int(match.group(1))
    match = re.match(r"n(\d+)_", scene_id)
    if match:
        return int(match.group(1))
    raise ValueError(f"Unable to infer n_objects from scene_id={scene_id!r}")


def _split_run_label(run_label: str) -> tuple[str, str]:
    parts = Path(run_label).parts
    if len(parts) >= 2:
        return parts[0], parts[-1]
    if parts:
        return parts[0], parts[0]
    raise ValueError(f"Invalid run label: {run_label!r}")


def _run_label_from_source_rel_path(rel_run: Path) -> str:
    parts = rel_run.parts
    if len(parts) >= 2:
        return f"{parts[0]}/{parts[-1]}"
    return str(rel_run)


def _candidate_run_labels_from_recon_rel_path(rel_run: Path) -> list[str]:
    parts = rel_run.parts
    candidates: list[str] = []
    if str(rel_run):
        candidates.append(str(rel_run))
    if len(parts) >= 2:
        candidates.append(f"{parts[0]}/{parts[-1]}")
        if "__" in parts[0]:
            source_attempt, _ = parts[0].split("__", 1)
            candidates.append(f"{source_attempt}/{parts[-1]}")
    if len(parts) == 1 and "__" in parts[0]:
        source_attempt, canonical = parts[0].split("__", 1)
        candidates.append(f"{source_attempt}/{canonical}")
    seen: set[str] = set()
    unique: list[str] = []
    for candidate in candidates:
        if candidate not in seen:
            seen.add(candidate)
            unique.append(candidate)
    return unique


def _sorted_scene_files(scene_dir: Path) -> list[Path]:
    return sorted(scene_dir.glob("*.json"))


def build_source_run_index(source_results_root: Path) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for run_dir in _discover_run_dirs(source_results_root):
        rel_run = run_dir.relative_to(source_results_root)
        run_label = _run_label_from_source_rel_path(rel_run)
        source_attempt, canonical_run = _split_run_label(run_label)
        summary = _load_json(run_dir / "summary.json")
        scene_index: dict[str, dict[str, Any]] = {}
        for scene_path in _sorted_scene_files(run_dir / "scenes"):
            scene_doc = _load_json(scene_path)
            scene_id = str(scene_doc.get("scene_id") or scene_path.stem)
            raw_model = scene_doc.get("model") or summary.get("model")
            scene_index[scene_id] = {
                "scene_id": scene_id,
                "n_objects": infer_n_objects(scene_id, scene_doc.get("n_objects")),
                "raw_model": raw_model,
                "source_scene_result": str(scene_path),
            }
        index[run_label] = {
            "run_label": run_label,
            "source_attempt": source_attempt,
            "canonical_run": canonical_run,
            "raw_model": summary.get("model"),
            "model_family": normalize_model_family(summary.get("model")),
            "summary_path": str(run_dir / "summary.json"),
            "scene_dir": str(run_dir / "scenes"),
            "n_scene_files": len(scene_index),
            "scene_index": scene_index,
        }
    return index


def _resolve_source_run_info(
    recon_root: Path,
    run_dir: Path,
    source_run_index: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    rel_run = run_dir.relative_to(recon_root)
    for candidate in _candidate_run_labels_from_recon_rel_path(rel_run):
        if candidate in source_run_index:
            return source_run_index[candidate]
    raise KeyError(
        f"Unable to map reconstruction run {rel_run} back to a source quick-sort run"
    )


def _numeric_mean(values: Iterable[Optional[float]]) -> Optional[float]:
    cleaned = [float(value) for value in values if value is not None]
    if not cleaned:
        return None
    return sum(cleaned) / len(cleaned)


def build_scene_long_rows(
    recon_root: Path,
    source_results_root: Path,
) -> list[dict[str, Any]]:
    source_run_index = build_source_run_index(source_results_root)
    rows: list[dict[str, Any]] = []
    for run_dir in _discover_recon_run_dirs(recon_root):
        source_run = _resolve_source_run_info(recon_root, run_dir, source_run_index)
        recon_summary = _load_json(run_dir / "recon" / "summary.json")
        recon_scene_dir = run_dir / "recon" / "scenes"
        for scene_path in _sorted_scene_files(recon_scene_dir):
            scene_doc = _load_json(scene_path)
            metrics = scene_doc.get("metrics", {})
            scene_id = str(scene_doc.get("scene_id") or scene_path.stem)
            raw_model = scene_doc.get("model") or source_run["raw_model"]
            rows.append(
                {
                    "model_family": normalize_model_family(raw_model),
                    "raw_model": raw_model,
                    "run_label": source_run["run_label"],
                    "source_attempt": source_run["source_attempt"],
                    "canonical_run": source_run["canonical_run"],
                    "scene_id": scene_id,
                    "n_objects": infer_n_objects(
                        scene_id,
                        scene_doc.get("n_objects")
                        or scene_doc.get("prepared_summary", {}).get("n_objects"),
                    ),
                    "status": scene_doc.get("status", "unknown"),
                    "feasible": bool(scene_doc.get("feasible")),
                    "is_skipped": False,
                    "skip_reason": None,
                    "source_scene_result": scene_doc.get("source_scene_result"),
                    "csr_qrr": metrics.get("csr_qrr"),
                    "csr_qrr_aligned": metrics.get("csr_qrr_aligned"),
                    "kendall_tau": metrics.get("kendall_tau"),
                    "nrms": metrics.get("nrms"),
                    "best_loss": metrics.get("best_loss"),
                    "n_solutions": metrics.get("n_solutions"),
                }
            )
        for skipped in recon_summary.get("skipped", []):
            scene_id = str(skipped["scene_id"])
            source_scene = source_run["scene_index"].get(scene_id, {})
            raw_model = source_scene.get("raw_model") or source_run["raw_model"]
            rows.append(
                {
                    "model_family": normalize_model_family(raw_model),
                    "raw_model": raw_model,
                    "run_label": source_run["run_label"],
                    "source_attempt": source_run["source_attempt"],
                    "canonical_run": source_run["canonical_run"],
                    "scene_id": scene_id,
                    "n_objects": infer_n_objects(
                        scene_id, source_scene.get("n_objects")
                    ),
                    "status": "skipped",
                    "feasible": False,
                    "is_skipped": True,
                    "skip_reason": skipped.get("reason"),
                    "source_scene_result": skipped.get("source_scene_result")
                    or source_scene.get("source_scene_result"),
                    "csr_qrr": None,
                    "csr_qrr_aligned": None,
                    "kendall_tau": None,
                    "nrms": None,
                    "best_loss": None,
                    "n_solutions": None,
                }
            )
    return _sort_scene_rows(rows)


def _sort_scene_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row["model_family"],
            row["run_label"],
            int(row["n_objects"]),
            row["scene_id"],
            row["status"],
        ),
    )


def _sort_group_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            row.get("model_family", ""),
            row.get("run_label", ""),
            int(row.get("n_objects", -1)),
        ),
    )


def aggregate_scene_rows(
    rows: list[dict[str, Any]],
    group_keys: Iterable[str],
    include_status_counts: bool = False,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    ordered_keys = tuple(group_keys)
    for row in rows:
        grouped[tuple(row[key] for key in ordered_keys)].append(row)

    aggregated_rows: list[dict[str, Any]] = []
    for key_values, group_rows in grouped.items():
        reconstructed_rows = [row for row in group_rows if not row["is_skipped"]]
        skipped_rows = [row for row in group_rows if row["is_skipped"]]
        status_counts = Counter(row["status"] for row in reconstructed_rows)
        aggregated = dict(zip(ordered_keys, key_values))
        aggregated["source_scenes"] = len(group_rows)
        aggregated["reconstructed"] = len(reconstructed_rows)
        aggregated["skipped"] = len(skipped_rows)
        aggregated["feasible"] = sum(1 for row in reconstructed_rows if row["feasible"])
        known_total = 0
        for status in KNOWN_STATUS_COLUMNS:
            count = status_counts.get(status, 0)
            aggregated[status] = count
            known_total += count
        aggregated["other_status"] = len(reconstructed_rows) - known_total
        aggregated["feasible_reconstructed"] = (
            aggregated["feasible"] / aggregated["reconstructed"]
            if aggregated["reconstructed"]
            else None
        )
        aggregated["feasible_total"] = (
            aggregated["feasible"] / aggregated["source_scenes"]
            if aggregated["source_scenes"]
            else None
        )
        for metric_name in MEAN_METRIC_COLUMNS:
            aggregated[f"{metric_name}_mean"] = _numeric_mean(
                row[metric_name] for row in reconstructed_rows
            )
        if include_status_counts:
            aggregated["status_counts"] = json.dumps(
                dict(sorted(status_counts.items())), ensure_ascii=False
            )
        aggregated_rows.append(aggregated)
    return _sort_group_rows(aggregated_rows)


def _get_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required to write the workbook. "
            "Run this script with `uv run --with openpyxl` or install openpyxl."
        ) from exc
    return Workbook, Alignment, Border, Font, PatternFill, Side


def _write_table_sheet(
    wb: Any,
    title: str,
    rows: list[dict[str, Any]],
    columns: Iterable[str],
    *,
    percent_columns: Iterable[str] = (),
) -> None:
    _, Alignment, Border, Font, PatternFill, Side = _get_openpyxl()
    thin = Border(*(Side(style="thin"),) * 4)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws = wb.create_sheet(title)
    headers = list(columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    percent_set = set(percent_columns)
    for row in rows:
        ws.append([row.get(column) for column in headers])

    for row in ws.iter_rows(min_row=2):
        for cell, column in zip(row, headers):
            cell.border = thin
            if column in percent_set and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
            elif column.endswith("_mean") and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _auto_width(ws: Any, lo: int = 8, hi: int = 30) -> None:
    for column in ws.columns:
        values = [len(str(cell.value or "")) for cell in column]
        width = min(max(values) + 3, hi) if values else lo
        ws.column_dimensions[column[0].column_letter].width = max(width, lo)


def write_workbook(
    output_path: Path,
    scene_rows: list[dict[str, Any]],
    pivot_model_run_n: list[dict[str, Any]],
    pivot_model_run: list[dict[str, Any]],
    pivot_model_n: list[dict[str, Any]],
) -> None:
    Workbook, _, _, _, _, _ = _get_openpyxl()
    wb = Workbook()
    wb.active.title = "_placeholder"
    percent_columns = ("feasible_reconstructed", "feasible_total")
    _write_table_sheet(wb, "scene_long", scene_rows, SCENE_LONG_COLUMNS)
    _write_table_sheet(
        wb,
        "pivot_model_run_n",
        pivot_model_run_n,
        PIVOT_MODEL_RUN_N_COLUMNS,
        percent_columns=percent_columns,
    )
    _write_table_sheet(
        wb,
        "pivot_model_run",
        pivot_model_run,
        PIVOT_MODEL_RUN_COLUMNS,
        percent_columns=percent_columns,
    )
    _write_table_sheet(
        wb,
        "pivot_model_n",
        pivot_model_n,
        PIVOT_MODEL_N_COLUMNS,
        percent_columns=percent_columns,
    )
    del wb["_placeholder"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def build_default_output_path(recon_root: Path) -> Path:
    return recon_root / "pivots" / "quicksort_bt0_recon_pivot.xlsx"


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build pivot-style Excel summaries for quick-sort BT=0 reconstruction outputs"
    )
    parser.add_argument(
        "--recon-root",
        required=True,
        help="Root directory containing per-run recon/summary.json outputs",
    )
    parser.add_argument(
        "--source-results-root",
        required=True,
        help="Root directory containing source quick-sort summary.json + scenes/ runs",
    )
    parser.add_argument(
        "--output-xlsx",
        default=None,
        help="Workbook output path; defaults to <recon-root>/pivots/quicksort_bt0_recon_pivot.xlsx",
    )
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    recon_root = Path(args.recon_root).resolve()
    source_results_root = Path(args.source_results_root).resolve()
    output_xlsx = (
        Path(args.output_xlsx).resolve()
        if args.output_xlsx
        else build_default_output_path(recon_root)
    )

    scene_rows = build_scene_long_rows(recon_root, source_results_root)
    pivot_model_run_n = aggregate_scene_rows(
        scene_rows, ("model_family", "run_label", "n_objects")
    )
    pivot_model_run = aggregate_scene_rows(
        scene_rows, ("model_family", "run_label"), include_status_counts=True
    )
    pivot_model_n = aggregate_scene_rows(scene_rows, ("model_family", "n_objects"))
    write_workbook(
        output_xlsx,
        scene_rows,
        pivot_model_run_n,
        pivot_model_run,
        pivot_model_n,
    )
    print(json.dumps({"output_xlsx": str(output_xlsx), "n_rows": len(scene_rows)}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
