#!/usr/bin/env python3
"""
生成综合分析 Excel：模型排行榜、视角对比、难度曲线、重建、逐场景明细。

用法：
    cd VLM-test
    python analysis/generate_insight_excel.py
    python analysis/generate_insight_excel.py --output path/to/output.xlsx
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).parent.parent))
from analysis.aggregate import load_scene_results

# ── 路径 ──
BASE_DIR = Path(__file__).parent.parent  # VLM-test/
RESULTS_DIR = BASE_DIR / "output" / "results"
ANALYSIS_DIR = BASE_DIR / "output" / "analysis"
RECON_DIR = ANALYSIS_DIR / "belief_recon"

# ── 样式 ──
_BLUE = "2F5496"
HDR_FILL = PatternFill(start_color=_BLUE, end_color=_BLUE, fill_type="solid")
HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WORST_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GAIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BASELINE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN = Border(*(Side(style="thin"),) * 4)
PCT = "0.0%"
F3 = "0.000"
DASH = "—"

# ── 模型显示名 ──
_FAMILY_MAP = {
    "claude-sonnet-4-6": "Claude Sonnet 4.6",
    "claude_opus": "Claude Opus 4",
    "doubao-seed-2_0-pro": "Doubao Seed 2.0 Pro",
    "gemini-3_1-pro-preview": "Gemini 3.1 Pro",
    "gpt_5_4": "GPT-5.4",
    "kimi_k2_5": "Kimi K2.5",
    "qwen3p5_397B": "Qwen3.5 397B",
}


def _family(name: str) -> str:
    for prefix, display in _FAMILY_MAP.items():
        if name.startswith(prefix):
            return display
    return name


def _view(name: str) -> str:
    return "multi" if "multi_view" in name else "single"


def _is_fdr_missing(summary: dict) -> bool:
    """FDR 全部 missing 的模型。"""
    o = summary.get("overall", {})
    return o.get("fdr_total", 0) > 0 and o.get("fdr_total", 0) == o.get("missing", 0)


# ── 数据加载 ──

def load_summaries() -> Dict[str, dict]:
    out = {}
    for d in sorted(RESULTS_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith("results-"):
            continue
        sp = d / "summary.json"
        if sp.exists():
            out[d.name] = json.load(open(sp))
    return out


def load_recon_agg() -> Dict[str, dict]:
    """按模型聚合重建 JSON。"""
    out = {}
    if not RECON_DIR.is_dir():
        return out
    for md in sorted(RECON_DIR.iterdir()):
        if not md.is_dir():
            continue
        files = list(md.glob("*.json"))
        if not files:
            continue
        vals = defaultdict(list)
        n_feasible = 0
        for f in files:
            d = json.load(open(f))
            if d.get("feasible"):
                n_feasible += 1
            m = d.get("metrics", {})
            for k in ("csr_qrr", "csr_trr", "nrms", "kendall_tau", "K_geom"):
                v = m.get(k)
                if isinstance(v, (int, float)):
                    vals[k].append(v)
        avg = lambda lst: sum(lst) / len(lst) if lst else None
        out[md.name] = {
            "n": len(files),
            "n_feasible": n_feasible,
            "feasible_rate": n_feasible / len(files) if files else 0,
            "csr_qrr": avg(vals["csr_qrr"]),
            "csr_trr": avg(vals["csr_trr"]),
            "nrms": avg(vals["nrms"]),
            "kendall_tau": avg(vals["kendall_tau"]),
            "K_geom": avg(vals["K_geom"]),
        }
    return out


def load_recon_by_split() -> Dict[str, Dict[str, dict]]:
    """按模型×split 聚合重建指标。返回 {model: {split: {...}}}。"""
    out: Dict[str, Dict[str, dict]] = {}
    if not RECON_DIR.is_dir():
        return out
    for md in sorted(RECON_DIR.iterdir()):
        if not md.is_dir():
            continue
        splits: Dict[str, dict] = defaultdict(lambda: {
            "n": 0, "feasible": 0, "has_nrms": 0,
            "csr_qrr": [], "csr_trr": [], "nrms": [], "kendall_tau": [],
        })
        for f in md.glob("*.json"):
            d = json.load(open(f))
            sid = d.get("scene_id", f.stem)
            sp = sid.rsplit("_", 1)[0]
            s = splits[sp]
            s["n"] += 1
            if d.get("feasible"):
                s["feasible"] += 1
            m = d.get("metrics", {})
            if isinstance(m.get("nrms"), (int, float)):
                s["has_nrms"] += 1
            for k in ("csr_qrr", "csr_trr", "nrms", "kendall_tau"):
                v = m.get(k)
                if isinstance(v, (int, float)):
                    s[k].append(v)
        # 转为均值
        result = {}
        for sp, s in splits.items():
            avg = lambda lst: sum(lst) / len(lst) if lst else None
            result[sp] = {
                "n": s["n"],
                "feasible": s["feasible"],
                "has_nrms": s["has_nrms"],
                "feasible_rate": s["feasible"] / s["n"] if s["n"] else None,
                "recon_rate": s["has_nrms"] / s["n"] if s["n"] else None,
                "csr_qrr": avg(s["csr_qrr"]),
                "nrms": avg(s["nrms"]),
                "kendall_tau": avg(s["kendall_tau"]),
            }
        out[md.name] = result
    return out


# ── 工具函数 ──

def _style_hdr(ws, row=1):
    for c in ws[row]:
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN


def _auto_width(ws, lo=8, hi=26):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, MergedCell)]
        if not cells:
            continue
        w = max(len(str(c.value or "")) for c in cells)
        ws.column_dimensions[cells[0].column_letter].width = max(lo, min(w + 3, hi))


def _highlight_col(ws, r_start, r_end, cols, higher=True):
    for ci in cols:
        vals = []
        for ri in range(r_start, r_end + 1):
            c = ws.cell(row=ri, column=ci)
            if isinstance(c.value, (int, float)):
                vals.append((ri, c.value))
        if len(vals) < 2:
            continue
        best = max(v for _, v in vals) if higher else min(v for _, v in vals)
        for ri, v in vals:
            if v == best:
                ws.cell(row=ri, column=ci).fill = BEST_FILL


def _border_all(ws, r_start, r_end):
    for row in ws.iter_rows(min_row=r_start, max_row=r_end):
        for c in row:
            c.border = THIN


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 1: Main Results
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_main(wb, summaries):
    ws = wb.active
    ws.title = "Main Results"

    headers = [
        "Rank", "Model", "View", "#Scenes",
        "QRR\nAcc", "TRR\nHour", "TRR\nQuad", "TRR\nAdj",
        "FDR\nExact", "FDR\nKendall", "FDR\nPairwise", "FDR\nTop1",
        "Composite", "Missing",
    ]
    ws.append(headers)
    _style_hdr(ws)
    ws.row_dimensions[1].height = 36

    rows = []
    for name, s in summaries.items():
        o = s.get("overall", {})
        fdr_miss = _is_fdr_missing(s)
        qrr = o.get("qrr_accuracy") or 0
        trr_h = o.get("trr_hour_accuracy") or 0
        trr_q = o.get("trr_quadrant_accuracy") or 0
        trr_a = o.get("trr_adjacent_accuracy") or 0
        fdr_e = None if fdr_miss else (o.get("fdr_exact_accuracy") or 0)
        fdr_k = None if fdr_miss else (o.get("fdr_kendall_mean") or 0)
        fdr_p = None if fdr_miss else (o.get("fdr_pairwise_mean") or 0)
        fdr_t = None if fdr_miss else (o.get("fdr_top1_mean") or 0)
        comp = (qrr + trr_h + (fdr_k or 0)) / 3
        rows.append((name, qrr, trr_h, trr_q, trr_a, fdr_e, fdr_k, fdr_p, fdr_t, comp, s.get("n_scenes", 0), o.get("missing", 0)))

    rows.sort(key=lambda r: r[9], reverse=True)

    # Random baseline
    RANDOM = {"qrr": 1 / 3, "trr_h": 1 / 12, "trr_q": 1 / 4, "trr_a": 3 / 12}

    data_start = 2
    for rank, (name, qrr, trr_h, trr_q, trr_a, fdr_e, fdr_k, fdr_p, fdr_t, comp, ns, miss) in enumerate(rows, 1):
        ws.append([
            rank, _family(name), _view(name), ns,
            qrr, trr_h, trr_q, trr_a,
            fdr_e if fdr_e is not None else DASH,
            fdr_k if fdr_k is not None else DASH,
            fdr_p if fdr_p is not None else DASH,
            fdr_t if fdr_t is not None else DASH,
            comp, miss,
        ])

    data_end = 1 + len(rows)

    # Random baseline row
    bl_row = data_end + 1
    ws.append([
        "", "Random Baseline", "—", "—",
        RANDOM["qrr"], RANDOM["trr_h"], RANDOM["trr_q"], RANDOM["trr_a"],
        DASH, DASH, DASH, DASH, DASH, "—",
    ])
    for c in ws[bl_row]:
        c.fill = BASELINE_FILL
        c.font = Font(italic=True, color="808080")

    # 格式化
    for row in ws.iter_rows(min_row=2, max_row=data_end, min_col=5, max_col=13):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = PCT
    for row in ws.iter_rows(min_row=2, max_row=data_end):
        for c in row:
            c.border = THIN
            c.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2, max_row=data_end, min_col=2, max_col=2):
        for c in row:
            c.alignment = Alignment(horizontal="left")

    _highlight_col(ws, data_start, data_end, list(range(5, 14)))
    _auto_width(ws)
    ws.freeze_panes = "A2"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 2: View Comparison
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_view(wb, summaries):
    ws = wb.create_sheet("View Comparison")

    headers = [
        "Model Family",
        "S_QRR", "M_QRR", "Δ QRR",
        "S_TRR_H", "M_TRR_H", "Δ TRR",
        "S_FDR_K", "M_FDR_K", "Δ FDR",
    ]
    ws.append(headers)
    _style_hdr(ws)

    families = defaultdict(dict)
    for name, s in summaries.items():
        fam = _family(name)
        v = _view(name)
        o = s.get("overall", {})
        fdr_miss = _is_fdr_missing(s)
        families[fam][v] = {
            "qrr": o.get("qrr_accuracy") or 0,
            "trr": o.get("trr_hour_accuracy") or 0,
            "fdr": None if fdr_miss else (o.get("fdr_kendall_mean") or 0),
        }

    data_start = 2
    for fam in sorted(families):
        sv = families[fam].get("single", {})
        mv = families[fam].get("multi", {})

        def _d(sk, mk):
            if sk is None or mk is None:
                return DASH
            return mk - sk

        row = [
            fam,
            sv.get("qrr"), mv.get("qrr"), _d(sv.get("qrr"), mv.get("qrr")),
            sv.get("trr"), mv.get("trr"), _d(sv.get("trr"), mv.get("trr")),
            sv.get("fdr") if sv.get("fdr") is not None else DASH,
            mv.get("fdr") if mv.get("fdr") is not None else DASH,
            _d(sv.get("fdr"), mv.get("fdr")),
        ]
        ws.append(row)

    data_end = ws.max_row

    delta_cols = [4, 7, 10]
    for row in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=2, max_col=10):
        for c in row:
            if isinstance(c.value, (int, float)):
                if c.column in delta_cols:
                    c.number_format = "+0.0%;-0.0%"
                    if c.value > 0.005:
                        c.fill = GAIN_FILL
                    elif c.value < -0.005:
                        c.fill = LOSS_FILL
                else:
                    c.number_format = PCT
            c.border = THIN
            c.alignment = Alignment(horizontal="center")

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 3: Difficulty Scaling
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_difficulty(wb, summaries):
    ws = wb.create_sheet("Difficulty (by N)")

    splits = ["n04", "n05", "n06", "n07", "n08", "n09", "n10"]

    # 每个模型族取 composite 最高的视角
    best_per_fam = {}
    for name, s in summaries.items():
        fam = _family(name)
        o = s.get("overall", {})
        fdr_k = o.get("fdr_kendall_mean") or 0
        if _is_fdr_missing(s):
            fdr_k = 0
        comp = ((o.get("qrr_accuracy") or 0) + (o.get("trr_hour_accuracy") or 0) + fdr_k) / 3
        if fam not in best_per_fam or comp > best_per_fam[fam][1]:
            best_per_fam[fam] = (name, comp)

    model_list = [(fam, best_per_fam[fam][0]) for fam in sorted(best_per_fam)]

    metrics = [
        ("QRR Accuracy", "qrr_accuracy"),
        ("TRR Hour Accuracy", "trr_hour_accuracy"),
        ("FDR Exact Accuracy", "fdr_exact_accuracy"),
        ("FDR Kendall τ", "fdr_kendall_mean"),
    ]

    cr = 1
    for label, key in metrics:
        ws.cell(row=cr, column=1, value=label).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=1 + len(splits))
        cr += 1

        hdr = ["Model"] + [s.upper() for s in splits]
        for ci, h in enumerate(hdr, 1):
            c = ws.cell(row=cr, column=ci, value=h)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = THIN
        cr += 1

        ds = cr
        for fam, dname in model_list:
            bs = summaries[dname].get("by_split", {})
            row = [fam]
            for sp in splits:
                v = bs.get(sp, {}).get(key)
                row.append(v)
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=cr, column=ci, value=v)
                if ci > 1 and isinstance(v, (int, float)):
                    c.number_format = PCT
                c.border = THIN
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            cr += 1

        _highlight_col(ws, ds, cr - 1, list(range(2, 2 + len(splits))))
        cr += 1

    _auto_width(ws)
    ws.freeze_panes = "B1"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 4: Reconstruction Summary
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_recon(wb, summaries, recon_agg):
    ws = wb.create_sheet("Reconstruction")

    headers = [
        "Model", "View", "#Scenes", "Feasible\nRate",
        "CSR\nQRR", "CSR\nTRR", "NRMS", "Kendall\nτ", "K_geom",
    ]
    ws.append(headers)
    _style_hdr(ws)
    ws.row_dimensions[1].height = 36

    data_start = 2
    for name in sorted(summaries):
        r = recon_agg.get(name, {})
        if not r:
            ws.append([_family(name), _view(name), DASH, DASH, DASH, DASH, DASH, DASH, DASH])
            continue
        ws.append([
            _family(name), _view(name), r["n"], r["feasible_rate"],
            r["csr_qrr"], r["csr_trr"], r["nrms"], r["kendall_tau"], r["K_geom"],
        ])

    data_end = ws.max_row

    for row in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=4, max_col=4):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = PCT
    for row in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=5, max_col=9):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = F3
    _border_all(ws, 1, data_end)
    for row in ws.iter_rows(min_row=data_start, max_row=data_end):
        for c in row:
            c.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=1, max_col=1):
        for c in row:
            c.alignment = Alignment(horizontal="left")

    # 高亮：CSR 越高越好，NRMS 越低越好
    _highlight_col(ws, data_start, data_end, [4, 5, 6, 8], higher=True)
    _highlight_col(ws, data_start, data_end, [7], higher=False)

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 4.5: Recon by N
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_recon_by_n(wb, summaries, recon_by_split):
    ws = wb.create_sheet("Recon by N")

    splits = ["n04", "n05", "n06", "n07", "n08", "n09", "n10"]

    # 全部 14 个模型，显示 model_family + view
    model_list = sorted(recon_by_split.keys())  # dir names

    # ── 计数子表：显示 "feasible/total" 格式 ──
    count_tables = [
        ("Feasible Count (feasible / total)", "feasible", "n"),
        ("Recon Count (has_positions / total)", "has_nrms", "n"),
    ]

    # ── 均值子表：显示数值 ──
    mean_tables = [
        ("CSR_QRR Mean", "csr_qrr", F3, True),
        ("NRMS Mean (↓ better)", "nrms", F3, False),
        ("Kendall τ Mean", "kendall_tau", F3, True),
    ]

    cr = 1

    # 先渲染计数子表
    for label, count_key, total_key in count_tables:
        ws.cell(row=cr, column=1, value=label).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=1 + len(splits))
        cr += 1

        hdr = ["Model"] + [s.upper() for s in splits]
        for ci, h in enumerate(hdr, 1):
            c = ws.cell(row=cr, column=ci, value=h)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = THIN
        cr += 1

        for dname in model_list:
            sp_data = recon_by_split.get(dname, {})
            row = [f"{_family(dname)} ({_view(dname)})"]
            for sp in splits:
                sd = sp_data.get(sp, {})
                cnt = sd.get(count_key, 0)
                tot = sd.get(total_key, 0)
                row.append(f"{cnt}/{tot}" if tot else DASH)
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=cr, column=ci, value=v)
                c.border = THIN
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            cr += 1
        cr += 1  # 空行

    # 再渲染均值子表
    for label, key, fmt, higher in mean_tables:
        ws.cell(row=cr, column=1, value=label).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=1 + len(splits))
        cr += 1

        hdr = ["Model"] + [s.upper() for s in splits]
        for ci, h in enumerate(hdr, 1):
            c = ws.cell(row=cr, column=ci, value=h)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = THIN
        cr += 1

        ds = cr
        for dname in model_list:
            sp_data = recon_by_split.get(dname, {})
            row = [f"{_family(dname)} ({_view(dname)})"]
            for sp in splits:
                v = sp_data.get(sp, {}).get(key)
                row.append(v)
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=cr, column=ci, value=v)
                if ci > 1 and isinstance(v, (int, float)):
                    c.number_format = fmt
                c.border = THIN
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            cr += 1

        _highlight_col(ws, ds, cr - 1, list(range(2, 2 + len(splits))), higher=higher)
        cr += 1  # 空行

    _auto_width(ws)
    ws.freeze_panes = "B1"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 5: Per Scene
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_per_scene(wb, summaries):
    ws = wb.create_sheet("Per Scene")

    headers = [
        "Model", "View", "Scene_ID", "Split", "N_Obj",
        "QRR_C", "QRR_T", "QRR%",
        "TRR_H_C", "TRR_Q_C", "TRR_A_C", "TRR_T",
        "TRR_H%", "TRR_Q%", "TRR_A%",
        "FDR_E_C", "FDR_T", "FDR_E%",
        "FDR_K", "FDR_P", "FDR_T1",
        "Miss",
    ]
    ws.append(headers)
    _style_hdr(ws)

    def _div(a, b):
        return a / b if b and b > 0 else None

    for name in sorted(summaries):
        scenes = load_scene_results(str(RESULTS_DIR / name))
        view = _view(name)
        for sr in scenes:
            sid = sr["scene_id"]
            sp = sid.rsplit("_", 1)[0]
            n = sr.get("n_objects", int(sid[1:3]))
            sc = sr["scores"]
            ws.append([
                _family(name), view, sid, sp, n,
                sc.get("qrr_correct", 0), sc.get("qrr_total", 0),
                _div(sc.get("qrr_correct", 0), sc.get("qrr_total", 0)),
                sc.get("trr_hour_correct", 0),
                sc.get("trr_quadrant_correct", 0),
                sc.get("trr_adjacent_correct", 0),
                sc.get("trr_total", 0),
                _div(sc.get("trr_hour_correct", 0), sc.get("trr_total", 0)),
                _div(sc.get("trr_quadrant_correct", 0), sc.get("trr_total", 0)),
                _div(sc.get("trr_adjacent_correct", 0), sc.get("trr_total", 0)),
                sc.get("fdr_exact_correct", 0), sc.get("fdr_total", 0),
                _div(sc.get("fdr_exact_correct", 0), sc.get("fdr_total", 0)),
                sc.get("fdr_kendall_mean"),
                sc.get("fdr_pairwise_mean"),
                sc.get("fdr_top1_mean"),
                sc.get("missing", 0),
            ])

    pct_cols = [8, 13, 14, 15, 18]
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.column in pct_cols and isinstance(c.value, (int, float)):
                c.number_format = PCT
            elif c.column in (19, 20, 21) and isinstance(c.value, (int, float)):
                c.number_format = F3

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Sheet 6: Key Insights
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sheet_insights(wb, summaries, recon_agg):
    ws = wb.create_sheet("Key Insights")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 90

    title_f = Font(bold=True, size=14, color=_BLUE)
    sec_f = Font(bold=True, size=11)
    body_f = Font(size=10)

    r = 1
    ws.cell(row=r, column=2, value="ORDINARY-BENCH 评估分析 (2026-03-30)").font = title_f
    r += 2

    # 汇总计算
    scores = {}
    for name, s in summaries.items():
        o = s.get("overall", {})
        fdr_k = o.get("fdr_kendall_mean") or 0
        if _is_fdr_missing(s):
            fdr_k = 0
        scores[name] = {
            "qrr": o.get("qrr_accuracy") or 0,
            "trr_h": o.get("trr_hour_accuracy") or 0,
            "trr_a": o.get("trr_adjacent_accuracy") or 0,
            "fdr_k": fdr_k,
            "fdr_e": 0 if _is_fdr_missing(s) else (o.get("fdr_exact_accuracy") or 0),
            "comp": ((o.get("qrr_accuracy") or 0) + (o.get("trr_hour_accuracy") or 0) + fdr_k) / 3,
            "ns": s.get("n_scenes", 0),
            "miss": o.get("missing", 0),
        }

    best = max(scores, key=lambda k: scores[k]["comp"])
    bs = scores[best]
    best_trr = max(scores, key=lambda k: scores[k]["trr_h"])
    best_fdr = max(scores, key=lambda k: scores[k]["fdr_k"])

    lines = []

    # 1 综合冠军
    lines.append(("1. 综合最佳", f"{_family(best)} ({_view(best)}) — Composite {bs['comp']:.1%}"))
    lines.append(("", f"   QRR {bs['qrr']:.1%} | TRR Hour {bs['trr_h']:.1%} | FDR Kendall {bs['fdr_k']:.1%}"))

    # 2 各项最佳
    lines.append(("2. 各项冠军", ""))
    lines.append(("", f"   QRR: {_family(max(scores, key=lambda k: scores[k]['qrr']))} — {scores[max(scores, key=lambda k: scores[k]['qrr'])]['qrr']:.1%}"))
    lines.append(("", f"   TRR Hour: {_family(best_trr)} — {scores[best_trr]['trr_h']:.1%}"))
    lines.append(("", f"   FDR Kendall: {_family(best_fdr)} — {scores[best_fdr]['fdr_k']:.1%}"))

    # 3 TRR 普遍短板
    near_random = [n for n, sc in scores.items() if sc["trr_h"] < 0.20]
    lines.append(("3. TRR 是普遍短板", f"{len(near_random)}/{len(scores)} 模型 TRR Hour < 20%（随机基线 8.3%）"))
    above = [n for n, sc in scores.items() if sc["trr_h"] > 0.40]
    if above:
        lines.append(("", f"   仅 {', '.join(_family(n) for n in above)} 显著超过随机"))

    # 4 Multi-view
    fam_views = defaultdict(dict)
    for n, sc in scores.items():
        fam_views[_family(n)][_view(n)] = sc
    diffs = []
    for fam, vs in fam_views.items():
        if "single" in vs and "multi" in vs:
            d = vs["multi"]["comp"] - vs["single"]["comp"]
            diffs.append((fam, d))
    if diffs:
        avg_d = sum(d for _, d in diffs) / len(diffs)
        lines.append(("4. Multi-View 效果", f"平均影响 {avg_d:+.1%} — 多视角总体{'有益' if avg_d > 0 else '有害'}"))
        for fam, d in sorted(diffs, key=lambda x: x[1]):
            lines.append(("", f"   {fam}: {d:+.1%}"))

    # 5 难度缩放
    bs_summary = summaries[best]
    bs_by = bs_summary.get("by_split", {})
    n04_q = bs_by.get("n04", {}).get("qrr_accuracy") or 0
    n10_q = bs_by.get("n10", {}).get("qrr_accuracy") or 0
    n04_f = bs_by.get("n04", {}).get("fdr_exact_accuracy") or 0
    n10_f = bs_by.get("n10", {}).get("fdr_exact_accuracy") or 0
    lines.append(("5. 难度缩放（最佳模型）", f"QRR: {n04_q:.1%}(n=4)→{n10_q:.1%}(n=10) | FDR Exact: {n04_f:.1%}(n=4)→{n10_f:.1%}(n=10)"))

    # 6 完成率
    max_ns = max(sc["ns"] for sc in scores.values())
    incomplete = [(n, sc["ns"]) for n, sc in scores.items() if sc["ns"] < max_ns * 0.9]
    fdr_miss_models = [n for n, s in summaries.items() if _is_fdr_missing(s)]
    if incomplete or fdr_miss_models:
        lines.append(("6. 数据完整性", ""))
        if incomplete:
            for n, ns in sorted(incomplete, key=lambda x: x[1]):
                lines.append(("", f"   {_family(n)} ({_view(n)}): {ns}/{max_ns} scenes"))
        if fdr_miss_models:
            lines.append(("", f"   FDR 全 missing: {', '.join(_family(n)+'('+_view(n)+')' for n in fdr_miss_models)}"))

    # 7 重建洞见
    if recon_agg:
        best_recon = max(recon_agg, key=lambda k: (recon_agg[k].get("kendall_tau") or 0))
        br = recon_agg[best_recon]
        lines.append(("7. 重建质量", f"最佳空间一致性: {_family(best_recon)} ({_view(best_recon)})"))
        lines.append(("", f"   CSR_QRR {br['csr_qrr']:.3f} | NRMS {br['nrms']:.3f} | Kendall τ {br['kendall_tau']:.3f}"))
        worst_recon = min(recon_agg, key=lambda k: (recon_agg[k].get("kendall_tau") or 999))
        wr = recon_agg[worst_recon]
        lines.append(("", f"   最差: {_family(worst_recon)} ({_view(worst_recon)}) — Kendall τ {wr['kendall_tau']:.3f}"))

    for sec, txt in lines:
        if sec:
            ws.cell(row=r, column=2, value=sec).font = sec_f
            r += 1
        if txt:
            ws.cell(row=r, column=2, value=txt).font = body_f
            r += 1
        if sec and sec[0].isdigit():
            r += 1

    ws.sheet_properties.tabColor = _BLUE


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Main
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    summaries = load_summaries()
    recon_agg = load_recon_agg()
    recon_splits = load_recon_by_split()
    print(f"Models: {len(summaries)}  |  Recon: {len(recon_agg)}  |  Recon splits: {len(recon_splits)}")

    wb = Workbook()
    sheet_main(wb, summaries)
    sheet_view(wb, summaries)
    sheet_difficulty(wb, summaries)
    sheet_recon(wb, summaries, recon_agg)
    sheet_recon_by_n(wb, summaries, recon_splits)
    sheet_per_scene(wb, summaries)
    sheet_insights(wb, summaries, recon_agg)

    out = Path(args.output) if args.output else ANALYSIS_DIR / "insight_report_0330.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
